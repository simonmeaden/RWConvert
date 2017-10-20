'''
Created on 5 Oct 2017

@author: Simon Meaden
'''
import logging
import sys
import os.path
import appdirs
from PyQt5.Qt import (
    QApplication, 
    QDesktopWidget,
    QMainWindow,
    QTabWidget,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QGroupBox,
    QFrame, 
    QPushButton, 
    QListWidget,
    QAbstractItemView,
    QComboBox,
    QCheckBox,
    QLabel,
    QLineEdit,
    QSizePolicy,
    QMessageBox
    )
from PyQt5.QtCore import pyqtSlot

from pluginbase import PluginBase
from functools import partial

from openpyxl import Workbook

import yaml
from yaml import YAMLObject 
from datetime import date, datetime
import const


# constant values
PATHS = 'Paths'
FILES = 'Files'
FORMS = 'Forms'
NAMES = 'Names'
 
class ConverterWidget(QMainWindow):
    
    name = 'RWConvert'
   
    # For easier usage calculate the path relative to here.
    here = os.path.abspath(os.path.dirname(__file__))
    getPath = partial(os.path.join, here)


    # Setup a plugin base for "rwconvert.plugins" and make sure to load
    # all the default built-in plugins from the builtin_plugins folder.
    pluginBase = PluginBase(package='rwconvert.plugins',
                            searchpath=[getPath('./plugins')])

    plugins = {}
    filenames = []
    currentPlugin = None
    filetypes = 'All Files (*.*)'
    config = {}
 
    def __init__(self):
        '''
        init
        '''
        
        super().__init__()
        
              
        """
        Read configuration file and return its contents
        """
        self.initConfig()
        cfgDir = self.config[PATHS]['config_path']
        self.cfgFileName = os.path.join(cfgDir, self.config[FILES]['config_name'])
        if not os.path.isfile(self.cfgFileName):
            os.makedirs(os.path.dirname(self.cfgFileName), exist_ok=True)
            self.exportConfig()
        else:
            with open(self.cfgFileName, 'r') as configFile:
                c = yaml.load(configFile)
                c = {} if c is None else c
                if c != {}:
                    self.config = c
                                            
        # and a source which loads the plugins from the "plugins"
        # folder.  We also pass the application name as identifier.  This
        # is optional but by doing this out plugins have consistent
        # internal module names which allows pickle to work.
        self.source = self.pluginBase.make_plugin_source(
            searchpath=[self.getPath('./plugins')],
            identifier=self.name)

        # Here we list all the plugins the source knows about, load them
        # and the use the "setup" function provided by the plugin to
        # initialize the plugin.
        for pluginName in self.source.list_plugins():
            plugin = self.source.load_plugin(pluginName)
            pluginClass = getattr(plugin, pluginName)
            instance = pluginClass()
            self.plugins[pluginName] = instance

        self.initGui()
#         self.showFullScreen()
        

    def initConfig(self):
        if PATHS not in self.config: self.config[PATHS] = {}
        if FILES not in self.config: self.config[FILES] = {}
        if FORMS not in self.config: self.config[FORMS] = {}
        if NAMES not in self.config: self.config[NAMES] = {}
        self.config[PATHS]['homepath']       = os.path.expanduser('~')
        self.config[PATHS]['docpath']        = os.path.join(self.config[PATHS]['homepath'], 'Documents')
        self.config[PATHS]['download_path']  = os.path.join(self.config[PATHS]['homepath'], 'Downloads')
        self.config[PATHS]['save_path']      = os.path.join(self.config[PATHS]['docpath'], self.name)
        self.config[PATHS]['config_path']    = appdirs.user_config_dir(self.name)
        self.config[FILES]['config_name']    = 'config.yaml'
        self.config[FILES]['save_file']      = 'roadwarrior'
        self.config[FILES]['save_ext']       = '.xlsx'
        self.config[FORMS]['include_date'] = False
        self.config[FORMS]['prefix_date'] = False
        self.config[FORMS]['include_routes'] = False
        self.config[FORMS]['combine_routes'] = False
        self.config[NAMES]['current_plugin_name'] = ''
        
        self.destFilename = self.config['Files']['save_file'] + self.config['Files']['save_ext']


    def initGui(self):
        self.setGeometry(300, 300, 800, 600)
        self.setWindowTitle('Road warrior Upload File Converter')
        self.center()
        
        fMain = QFrame()
        mainLayout = QGridLayout()
        fMain.setLayout(mainLayout)
        self.setCentralWidget(fMain)

        tabWidget = QTabWidget()
        mainLayout.addWidget(tabWidget, 0, 0)
#         mainLayout.addWidget(self.initConvertPage(), 0, 0)
        
        self.closeBtn = QPushButton('Close Application')
        self.closeBtn.clicked.connect(self.handleCloseClicked)
        mainLayout.addWidget(self.closeBtn, 1, 0)
         
        tabWidget.addTab(self.initConvertPage(), 'Converter')
        tabWidget.addTab(self.initConfigPage(), 'Configuration')
        
         
    def initConvertPage(self):
        f = QFrame()
        l = QGridLayout()
        f.setLayout(l)

        row = 0
        
        l.addWidget(QLabel('Converter :'), row, 0)
        currentPluginBox = QComboBox()
        currentPluginBox.currentTextChanged.connect(self.selectPlugin)
        l.addWidget(currentPluginBox, row, 0, 1, 2)
        for key in self.plugins.keys():                    
            currentPluginBox.addItem(key)
        row += 1
        
        destLbl = QLabel('Destination path :')
        destLbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        l.addWidget(destLbl, row, 0)
        self.destPathLbl = QLabel(self.joinSavePath(self.config[PATHS]['save_path'], 
                                                    self.config[FILES]['save_file'], 
                                                    self.config[FILES]['save_ext']))
        self.destPathLbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        l.addWidget(self.destPathLbl, row, 1)

        row += 1
        
        f2 = QFrame()
        l2 = QHBoxLayout()
        f2.setLayout(l2)
        l2.addWidget(self.initFileFrame())
        l2.addWidget(self.initOutputSetup())
        l.addWidget(f2, row, 0, 1, 2)
        
        row += 1
        
        l.addWidget(self.initSrcFiles(), row, 0, 1, 2)
        
        row += 1
            
        self.convertBtn = QPushButton('Convert')
        self.convertBtn.clicked.connect(self.handleConvertClicked)
        self.convertBtn.setEnabled(False)
        l.addWidget(self.convertBtn, row, 0, 1, 2)

        return f
    
    def initOutputSetup(self):
        row = 0
        
        f = QFrame()
        l = QGridLayout()
        f.setLayout(l)
#         f.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        
        l.addWidget(QLabel("Empty label"), 0, 0)
        
        return f
    
    def initSrcFiles(self):
        
        row = 0
        
        f = QFrame()
        l = QGridLayout()
        f.setLayout(l)

        l.addWidget(QLabel('Source Files :'), row, 0)
        self.srcFilesList = QListWidget()
        self.srcFilesList.setSelectionMode(QAbstractItemView.MultiSelection)
        self.srcFilesList.setAlternatingRowColors(True)
        self.srcFilesList.model().rowsInserted.connect(self.handleSrcFilesChanged)
        self.srcFilesList.model().rowsRemoved.connect(self.handleSrcFilesChanged)
        selectionModel = self.srcFilesList.selectionModel()
        selectionModel.selectionChanged.connect(self.handleSrcFilesSelectionChanged)
        l.addWidget(self.srcFilesList, row, 1, 3, 1)
        
        self.srcFilesBtn = QPushButton('Select Files')
        self.srcFilesBtn.clicked.connect(self.handleSelectSrcFiles)
        self.srcFilesBtn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        l.addWidget(self.srcFilesBtn, row, 2)
        row += 1
            
        self.addFilesBtn = QPushButton('Add Files')
        self.addFilesBtn.clicked.connect(self.handleAddSrcFiles)
        self.addFilesBtn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        l.addWidget(self.addFilesBtn, row, 2)
        row += 1
            
        self.removeFilesBtn = QPushButton('Remove Files')
        self.removeFilesBtn.setEnabled(False)
        self.removeFilesBtn.clicked.connect(self.handleRemoveSrcFiles)
        self.removeFilesBtn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        l.addWidget(self.removeFilesBtn, row, 2)
        
        return f
    
    def initFileFrame(self):
        row = 0
        
        f = QFrame()
        l = QGridLayout()
        f.setLayout(l)
#         f.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        
        l.addWidget(QLabel('Destination File :'), row, 0)
        self.destFilenameEdit = QLineEdit(self.destFilename)
        self.destFilenameEdit.setEnabled(False)
        self.destFilenameEdit.textChanged.connect(self.handleDestFilenameChanged)
        l.addWidget(self.destFilenameEdit, row, 1, 1, 3)
        row += 1
        
        self.combineRoutesBox = QCheckBox('combine_routes')
        if self.config[FORMS]['combine_routes']:
            self.combineRoutesBox.setChecked(True)
        else:
            self.combineRoutesBox.setChecked(False)
        self.combineRoutesBox.clicked.connect(self.handleCombineRoutesClicked)
        self.combineRoutesBox.setEnabled(False)
        l.addWidget(self.combineRoutesBox, row, 0, 1, 3)
        row += 1
        
        addDateInNameBox = QCheckBox('Add date to filename')
        addDateInNameBox.clicked.connect(self.handleAddDateClicked)
        l.addWidget(addDateInNameBox, row, 0, 1, 3)
        if self.config[FORMS]['prefix_date']:
            self.prefixDateInNameBox = QPushButton('prefix_date')
            self.prefixDateInNameBox.setEnabled(False)
        else:
            self.prefixDateInNameBox = QPushButton('Suffix date')
            self.prefixDateInNameBox.setEnabled(False)
        self.prefixDateInNameBox.setToolTip('Click to change to prefix date.')
        self.prefixDateInNameBox.clicked.connect(self.handlePrefixDateClicked)

        l.addWidget(self.prefixDateInNameBox, row, 1, 1, 2)
        row += 1
        
        addRouteInNameBox = QCheckBox('Add route to filename')
        addRouteInNameBox.clicked.connect(self.handleAddRouteClicked)
        l.addWidget(addRouteInNameBox, row, 0)
#         row += 1
                                        
        return f
        
    
    def initConfigPage(self):
        row = 0
        
        f = QFrame()
        l = QGridLayout()
        f.setLayout(l)
        
        srcLbl = QLabel('Source directory :')
        srcLbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        l.addWidget(srcLbl, row, 0)
        
        self.srcDirBox = QLineEdit()
        self.srcDirBox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.srcDirBox.setText(self.config[PATHS]['download_path'])
        l.addWidget(self.srcDirBox, row, 1)
        
        srcDirSelectBtn = QPushButton('Select')
        srcDirSelectBtn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        srcDirSelectBtn.clicked.connect(self.handleSelectSrcDirectory)
        l.addWidget(srcDirSelectBtn, row, 2)
        
        row += 1
    
        destLbl = QLabel('Destination directory :')
        destLbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        l.addWidget(destLbl, row, 0)
        
        self.destDirBox = QLineEdit()
        self.destDirBox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.destDirBox.setText(self.config[PATHS]['save_path'])
        l.addWidget(self.destDirBox, row, 1)
        
        destDirSelectBtn = QPushButton('Select')
        destDirSelectBtn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        destDirSelectBtn.clicked.connect(self.handleSelectDestDirectory)
        l.addWidget(destDirSelectBtn, row, 2)
        
        row += 1
         
        l.addWidget(QFrame(), row, 0, 1, 3)
        
        return f
    
    def handleSelectSrcDirectory(self):                                      
        directory = QFileDialog.getExistingDirectory(self,
                                                     'Select Source Directory', 
                                                     self.config[PATHS]['download_path'], 
                                                     QFileDialog.ShowDirsOnly);  
        if not directory == '':
            self.config[PATHS]['download_path'] = directory
            self.srcDirBox.setText(directory)

        
    def handleSelectDestDirectory(self):
        directory = QFileDialog.getExistingDirectory(self,
                                                     'Select Destination Directory', 
                                                     self.config[PATHS]['save_path'], 
                                                     QFileDialog.ShowDirsOnly);  
        if not directory == '':
            self.config[PATHS]['save_path'] = directory
            self.destDirBox.setText(directory)
            
    def joinSavePath(self, path, filename, ext):
        newpath = os.path.join(path, filename)
        newpath += ext
        return newpath

    @pyqtSlot(bool)
    def handleAddDateClicked(self, checked):
        if checked:
            self.config[FORMS]['include_path'] = True
            self.prefixDateInNameBox.setEnabled(True)
        else:
            self.config[FORMS]['include_path'] = False
            self.prefixDateInNameBox.setEnabled(False)
            
    @pyqtSlot()
    def handlePrefixDateClicked(self):
        if self.config[FORMS]['prefix_date']:
            self.config[FORMS]['prefix_date'] = False
            self.prefixDateInNameBox.setText('Suffix date')
            self.prefixDateInNameBox.setToolTip('Click to change to prefix date.')
        else:
            self.config[FORMS]['prefix_date'] = True
            self.prefixDateInNameBox.setText('Prefix date')
            self.prefixDateInNameBox.setToolTip('Click to change to suffix date.')

    @pyqtSlot()
    def handleAddRouteClicked(self, checked):
        if checked:
            self.config[FORMS]['include_routes'] = True
        else:
            self.config[FORMS]['include_routes'] = False
                        
    @pyqtSlot()
    def handleCombineRoutesClicked(self, checked):
        if checked:
            self.config[FORMS]['combine_routes'] = True
        else:
            self.config[FORMS]['combine_routes'] = False
            
        self.enableStuff()
            
    def enableStuff(self):     
        if self.srcFilesList.model().rowCount() == 0:
            self.convertBtn.setEnabled(False)
            self.combineRoutesBox.setEnabled(False)
            self.destFilenameEdit.setEnabled(False)
        elif self.srcFilesList.model().rowCount() == 1:
            self.convertBtn.setEnabled(True)
            self.combineRoutesBox.setEnabled(False)
            self.destFilenameEdit.setEnabled(True)
        else:
            self.convertBtn.setEnabled(True)
            self.combineRoutesBox.setEnabled(True)
            if self.combineRoutesBox.isChecked():
                self.destFilenameEdit.setEnabled(True)
            else:
                self.destFilenameEdit.setEnabled(False)
            
    @pyqtSlot()
    def handleDestFilenameChanged(self, text):        
        if len(text) > 0:
            if self.config[PATHS]['include_path']:
                year = datetime.year
                month = datetime.month
                day = datetime.day
                datestr = str(year)
                if month < 10: datestr += '0'
                datestr += str(month)
                if day < 10: datestr += '0'
                datestr += str(day)
                if self.config[FORMS]['prefix_date']:
                    name = datestr + '_' + text
                else :
                    name = text + '_' + datestr
            else:
                name = text
                       
            self.config[FILES]['save file'] = name
            self.destPathLbl.setText(self.joinSavePath(self.savepath, self.savefile, self.saveext))
          
    def createDestinationFilePath(self):
        return self.joinSavePath(self.savepath, self.savefile, self.saveext)
            
    @pyqtSlot()
    def handleSrcFilesSelectionChanged(self, selected):
        if len(selected.indexes()) > 0:
            self.removeFilesBtn.setEnabled(True)
        else:
            self.removeFilesBtn.setEnabled(False)
              
    @pyqtSlot()
    def handleSrcFilesChanged(self):
        self.enableStuff()
        
    @pyqtSlot()
    def handleConvertClicked(self):
        if len(self.filenames) > 0:
            toexcel = ToExcel()
     
            self.currentPlugin.convert(self.filenames)
            routedata = self.currentPlugin.rwdata
            
            if self.config[FORMS]['combine_routes']:
                combined = []
                for route in routedata.keys():
                    l = routedata[route]
                    combined += l
                toexcel.create_workbook(combined, self.joinSavePath(self.savepath, self.savefile, self.saveext))
                
            else:
                i = 1
                for route in routedata.keys():
                    l = routedata[route]
                    if self.includeRoutes and len(route) > 0:
                        self.joinSavePath(self.savepath, self.savefile + '_' + route, self.saveext)
                    else:
                        self.joinSavePath(self.savepath, self.savefile + '(' + str(i) + ')', self.saveext)
                        i += 1
                        
                    toexcel.create_workbook(combined, self.savefile)        
        
    @pyqtSlot()
    def handleConverterChanged(self):
        pluginName = self.currentPluginBox.currentText()
        if pluginName != self.config[NAMES]['current_plugin_name']:
            self.config[NAMES]['current_plugin_name'] = pluginName
            self.currentPlugin = self.plugins[pluginName]
            self.filetypes = self.currentPlugin.filetypes

    @pyqtSlot()
    def handleCloseClicked(self):
        self.close()
        
    @pyqtSlot()
    def handleSelectSrcFiles(self):
        fileDlg = QFileDialog(self, 'Select Files', self.downloadpath, self.filetypes)
        fileDlg.setFileMode(QFileDialog.ExistingFiles)

        if fileDlg.exec_():
            self.filenames = fileDlg.selectedFiles()
            
        self.srcFilesList.clear()
        self.srcFilesList.addItems(self.filenames)
        
    @pyqtSlot()
    def handleAddSrcFiles(self):
        fileDlg = QFileDialog(self, 'Select Files', self.downloadpath, self.filetypes)
        fileDlg.setFileMode(QFileDialog.ExistingFiles)

        if fileDlg.exec_():
            for filename in fileDlg.selectedFiles():
                if filename not in self.filenames:
                    self.filenames.append(filename)
                    self.srcFilesList.addItem(filename)
            
    @pyqtSlot()
    def handleRemoveSrcFiles(self):
        selectedModel = self.srcFilesList.selectionModel()
        selected = selectedModel.selectedIndexes()
        for index in selected:
            name = index.data()
            self.filenames.remove(name)
            self.srcFilesList.takeItem(index.row())
        selectedModel.clear()
        
    @pyqtSlot(str)
    def selectPlugin(self, name):
            # activate the current plugin
#             self.pluginManager.activatePluginByName(name)
            self.config[NAMES]['current_plugin_name'] = name
            self.currentPlugin = self.plugins[name]
            self.filetypes = self.currentPlugin.filetypes
            
    def center(self):    
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        
   
    def exportConfig(self):
        with open(self.cfgFileName, 'w') as configFile:
            yaml.dump(self.config, configFile)
            
    def createUserConfig(self):
        """
        Create the user's config file in OS specific location
        """
        os.makedirs(os.path.dirname(self.cfgFileName), exist_ok=True)
        self.exportConfig()
            
    def closeEvent(self, event):
        buttonReply = QMessageBox.warning(self, 
                                          'Close Application', 
                                          'You are about to close the Application,'
                                          ' Press Yes to continue or Cancel to return to the Application',
                                           QMessageBox.Yes | QMessageBox.Cancel, QMessageBox.Cancel)
        if buttonReply == QMessageBox.Yes:
            self.exportConfig()
            exit()
 
class ToExcel:
    
    def __init__(self):
        '''
        '''

    def create_workbook(self, data, savefile):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['B1'] = 'Street'
        ws['C1'] = 'City'
        ws['D1'] = 'State/Region'
        ws['E1'] = 'Postal Code'
        ws['F1'] = 'Country Code'
        ws['G1'] = 'Priority (0-1)'
        ws['H1'] = 'Phone'
        ws['I1'] = 'Note'
        ws['J1'] = 'Latitude'
        ws['K1'] = 'Longitude'
        ws['L1'] = 'Service Time'
        col_list = 'ABCDEFGHIJKL'
        for row in range(0,  len(self.data) - 1):
            for col in range(0, 9):
                cell = col_list[col] + str(row + 2)   
                c = data[row][col]     
                ws[cell] = c
        
        wb.save(savefile)


logging.basicConfig(level=logging.DEBUG)

if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    
    w = ConverterWidget()
    w.show()
    
    sys.exit(app.exec_())
